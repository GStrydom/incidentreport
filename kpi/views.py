from django.shortcuts import render, redirect, get_object_or_404
from .models import KPIIncidentReport, KPILead, KPITeam, KPICompany
from .forms import IReportForm
from django.contrib.auth.decorators import login_required

import openpyxl
import datetime


@login_required(login_url='/')
def kpi_home_page(request):
    context = dict()
    context['incidents'] = KPIIncidentReport.objects.all()
    return render(request, 'templates/index.html', context)


@login_required(login_url='/')
def incident_detail(request, incident_id):
    context = dict()
    context['incident'] = KPIIncidentReport.objects.get(pk=incident_id)
    return render(request, 'templates/incident_detail.html', context)


@login_required(login_url='/')
def kpi_leads(request):
    context = dict()
    context['companies'] = KPICompany.objects.all()
    context['leads'] = KPILead.objects.all()
    return render(request, 'templates/eftleads.html', context)


@login_required(login_url='/')
def kpi_teams(request):
    context = dict()
    context['companies'] = KPICompany.objects.all()
    context['teams'] = KPITeam.objects.all()
    return render(request, 'templates/eftteams.html', context)


@login_required(login_url='/')
def eftlead_detail(request, eftlead_id):
    context = dict()
    context['eftlead'] = KPILead.objects.get(pk=eftlead_id)
    context['team_names'] = KPITeam.objects.filter(eft_lead__id=eftlead_id)
    context['incidents'] = KPIIncidentReport.objects.filter(eft_lead__id=eftlead_id)
    context['image_url'] = context['eftlead'].profile_pic.url
    return render(request, 'templates/eftlead.html', context)


@login_required(login_url='/')
def eftteam_detail(request, eftteam_id):
    context = dict()
    context['eftteam'] = KPITeam.objects.get(pk=eftteam_id)
    context['eftleads'] = KPILead.objects.filter(kpiteam__id=eftteam_id)
    context['incidents'] = KPIIncidentReport.objects.filter(team_name__id=eftteam_id)
    context['team_picture'] = context['eftteam'].team_pic.url
    return render(request, 'templates/eftteam.html', context)


@login_required(login_url='/')
def kpi_company(request, company_id):
    context = dict()
    context['company'] = KPICompany.objects.get(pk=company_id)

    context['eftleads'] = KPILead.objects.filter(company__id=company_id)
    context['lead_count'] = len(KPILead.objects.filter(company__id=company_id))

    context['teams'] = KPITeam.objects.filter(company__id=company_id)
    context['team_count'] = len(KPITeam.objects.filter(company__id=company_id))

    context['incidents'] = KPIIncidentReport.objects.filter(company__id=company_id)
    context['inc_count'] = len(KPIIncidentReport.objects.filter(company__id=company_id))

    context['company_picture'] = context['company'].company_pic.url
    return render(request, 'templates/eft_company.html', context)


@login_required(login_url='/')
def kpi_create(request):
    if request.method == "POST":
        form = IReportForm(request.POST)
        if form.is_valid():
            incident = KPIIncidentReport()
            incident.date = form.cleaned_data['date']
            incident.ir_num = form.cleaned_data['ir_num']

            incident.company = form.cleaned_data['company']
            incident.eft_lead = form.cleaned_data['eft_lead']
            incident.team_name = form.cleaned_data['team_name']
            incident.incident = form.cleaned_data['incident']
            incident.description = form.cleaned_data['description']
            incident.hours_deducted = form.cleaned_data['hours_deducted']
            incident.reportable = form.cleaned_data['reportable']
            incident.reason = form.cleaned_data['reason']

            incident.save()
            return redirect('kpi_home_page')
        else:
            print("Form is not valid.")
    else:
        form = IReportForm()
    return render(request, 'templates/eftcreate.html', {'form': form})


@login_required(login_url='/')
def kpi_update(request, incident_id):
    incident = get_object_or_404(KPIIncidentReport, pk=incident_id)
    form = IReportForm(request.POST or None, instance=incident)
    if form.is_valid():
        instance = form.save(commit=False)
        instance.save()
        return redirect('kpi_home_page')
    return render(request, 'templates/eftcreate.html', {'form': form})


@login_required(login_url='/')
def kpi_export(request, incident_id):
    incident = get_object_or_404(KPIIncidentReport, pk=incident_id)
    wb = openpyxl.load_workbook("Incident Report.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")
    ws.cell(row=1, column=1).value = "Incident Report " + incident.ir_num
    ws.cell(row=3, column=5).value = incident.date
    ws.cell(row=4, column=5).value = incident.ir_num
    ws.cell(row=5, column=5).value = incident.company.name
    ws.cell(row=6, column=5).value = incident.eft_lead.name
    ws.cell(row=7, column=5).value = incident.team_name.name
    ws.cell(row=8, column=5).value = incident.hours_deducted
    ws.cell(row=9, column=5).value = incident.reportable
    ws.cell(row=10, column=5).value = incident.reason
    ws.cell(row=12, column=5).value = datetime.datetime.now()
    ws.cell(row=13, column=5).value = ''

    file_name = "Incident Report " + incident.ir_num + ".xlsx"

    wb.save(file_name)
    return redirect('kpi_home_page')


@login_required(login_url='/')
def kpi_delete(request, incident_id):
    incident = get_object_or_404(KPIIncidentReport, pk=incident_id)
    incident.delete()
    return redirect('kpi_home_page')


@login_required(login_url='/')
def kpi_duplicate(request):
    return render(request, 'templates/kpi_duplicate.html')


@login_required(login_url='/')
def kpi_help(request):
    context = dict()
    return render(request, 'templates/help.html', context)